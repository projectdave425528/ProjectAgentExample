"""Generate Dave transactions Excel report."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "交易記錄"

headers = ["日期", "客戶名稱", "維修項目", "數量", "報價金額（單價）", "實收金額", "付款方式", "付款狀態", "備註"]
header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
header_font = Font(bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

data = [
    ["2019-01-17", "Dave", "Screens swap settlement", 1, "100.00", "", "未知", "未付", "Dave計算欠$100"],
    ["2019-04-03", "Dave", "XR LCD", 2, "1050.00", "", "未知", "未付", ""],
    ["2019-04-03", "Dave", "XR Swap", 1, "500.00", "", "未知", "未付", ""],
    ["2019-04-06", "Dave", "i6 repair", 1, "250.00", "", "未知", "未付", ""],
    ["2019-04-06", "Dave", "i6s LCD", 2, "", "", "未知", "未付", "價格未明確"],
    ["2019-04-08", "Dave", "XR screen protector", 10, "12.00", "", "未知", "未付", ""],
    ["2019-04-09", "Dave", "XR LCD", 2, "1050.00", "", "未知", "未付", ""],
    ["2019-04-09", "Dave", "i6 battery", 10, "50.00", "", "未知", "未付", ""],
    ["2019-04-25", "Dave", "綜合訂單(screens+batteries+repair)", 1, "2960.00", "2960.00", "銀行轉帳", "已付", "含6s backlight@80 + 7+ home button@50"],
    ["2019-05-02", "Dave", "XS LCD", 2, "2000.00", "", "未知", "未付", ""],
    ["2019-05-05", "Dave", "XR screen protector", 20, "10.00", "", "未知", "未付", "min order $200"],
    ["2019-05-07", "Dave", "iX LCD", 1, "1700.00", "", "未知", "未付", "減價後"],
    ["2019-05-07", "Dave", "XS LCD", 1, "2000.00", "", "未知", "未付", ""],
    ["2019-05-07", "Dave", "Screen protectors (balance)", 1, "100.00", "", "未知", "未付", "上次付$100,補$100"],
    ["2019-05-24", "Dave", "i8 Touch ID repair", 1, "400.00", "", "未知", "未付", ""],
    ["2019-07-08", "Dave", "7+ Touch ID repair", 1, "", "", "未知", "未付", "master修好,價格待確認"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Summary row
summary_row = len(data) + 2
ws.cell(row=summary_row, column=1, value="總計")
ws.cell(row=summary_row, column=1).font = Font(bold=True)

quoted_total = sum(float(r[4]) * r[3] for r in data if r[4])
ws.cell(row=summary_row, column=5, value=f"{quoted_total:.2f}")

received_total = sum(float(r[5]) for r in data if r[5])
ws.cell(row=summary_row, column=6, value=f"{received_total:.2f}")

# Auto-adjust column width
for col_idx in range(1, len(headers) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = 10
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) + 2)
    ws.column_dimensions[col_letter].width = min(max_len, 35)

output_path = "output/dave_transactions.xlsx"
wb.save(output_path)
print(f"Done: {output_path}")
print(f"Records: {len(data)}")
print(f"Quoted total (qty x unit): ${quoted_total:.2f}")
print(f"Received total: ${received_total:.2f}")
