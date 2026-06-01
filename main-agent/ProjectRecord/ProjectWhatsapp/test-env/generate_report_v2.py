"""Generate Dave transactions Excel report v2 - full chat analysis."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "交易記錄"

headers = ["日期", "客戶名稱", "維修項目", "數量", "報價金額（單價）", "實收金額", "付款方式", "付款狀態", "備註"]
header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
header_font = Font(bold=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# 整理自完整對話 (2019/01 - 2020/03)，客戶: Dave
# 報價來自 Scofield 阿黎 的報價訊息；實收來自 "paid" 確認
data = [
    # 日期, 客戶, 項目, 數量, 報價單價, 實收, 付款方式, 狀態, 備註
    ["2019-01-17", "Dave", "Screens swap 結算", 1, "100.00", "100.00", "未知", "已付", "Dave自己計算欠$100"],
    ["2019-04-03", "Dave", "XR LCD", 2, "1050.00", "", "未知", "未付", "XR LCD@1050"],
    ["2019-04-03", "Dave", "XR Swap", 1, "500.00", "", "未知", "未付", "Swap@500"],
    ["2019-04-06", "Dave", "i6 維修", 1, "250.00", "", "未知", "未付", "i6 fix@250"],
    ["2019-04-08", "Dave", "XR screen protector", 10, "12.00", "", "未知", "未付", "@$12/件"],
    ["2019-04-09", "Dave", "i6 battery", 10, "50.00", "", "未知", "未付", "@$50/件"],
    ["2019-04-25", "Dave", "綜合訂單(screens+battery+repair)", 1, "2960.00", "2960.00", "銀行轉帳", "已付", "含6s backlight@80 + 7+ home button@50；先付$2830後補至$2960"],
    ["2019-05-02", "Dave", "XS LCD", 2, "2000.00", "", "未知", "未付", "XS LCD@2000"],
    ["2019-05-05", "Dave", "XR/iX screen protector", 20, "10.00", "", "未知", "未付", "20pcs總$200, min order"],
    ["2019-05-07", "Dave", "iX LCD", 1, "1700.00", "", "未知", "未付", "減價後$1700"],
    ["2019-05-24", "Dave", "i8 Touch ID 維修", 1, "400.00", "", "未知", "未付", "Master fix@400"],
    ["2019-09-03", "Dave", "i7 audio 維修", 1, "", "", "未知", "未付", "iPhone 7 audio problem unpaid"],
    ["2019-09-14", "Dave", "7+ 4013 維修", 1, "400.00", "", "未知", "未付", "7+ 4013 fix@$400"],
    ["2019-09-24", "Dave", "XS 維修", 1, "600.00", "", "未知", "未付", "Master fix@$600"],
    ["2019-10-03", "Dave", "XS 後鏡頭玻璃維修", 1, "100.00", "", "未知", "未付", "camera glass fix@100；連XS維修共$700"],
    ["2019-10-17", "Dave", "XS Max Swap", 1, "500.00", "", "未知", "未付", "Xsmax glass swap@500"],
    ["2019-10-17", "Dave", "XS Max LCD", 1, "1950.00", "", "未知", "未付", "Xsmax LCD@1950"],
    ["2019-10-22", "Dave", "iX Swap (man-made)", 1, "400.00", "", "未知", "未付", "boss同意swap@400"],
    ["2019-11-03", "Dave", "iX glass protector", 20, "20.00", "", "未知", "未付", "加價後@$20/件(iPhone 11版)"],
    ["2019-11-09", "Dave", "XS LCD", 1, "1400.00", "", "未知", "未付", "XS降價@1400"],
    ["2019-11-29", "Dave", "綜合訂單(seal+XR back plate+screens)", 1, "6940.00", "6940.00", "銀行轉帳", "已付", "Dave確認$6940並轉帳"],
    ["2019-12-11", "Dave", "i7 加速感應器維修", 1, "350.00", "", "未知", "未付", "fix@350(指南針/旋轉功能)"],
    ["2019-12-11", "Dave", "iX LCD", 1, "1200.00", "", "未知", "未付", "Master said 1200"],
    ["2020-01-13", "Dave", "i7 加速感應器維修", 1, "350.00", "", "未知", "未付", "議價400→350"],
    ["2020-02-22", "Dave", "i11 LCD", 2, "700.00", "", "未知", "未付", "i11@700(LCD版)"],
    ["2020-02-24", "Dave", "7+ 維修", 1, "400.00", "", "未知", "未付", "7plus fix@400"],
    ["2020-02-29", "Dave", "i7 主板維修", 1, "350.00", "", "未知", "未付", "i7 motherboard fix@350"],
    ["2020-03-08", "Dave", "綜合訂單(screens+repair)", 1, "2940.00", "2940.00", "銀行轉帳", "已付", "Dave確認$2940並轉帳俾boss"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Summary row
summary_row = len(data) + 2
ws.cell(row=summary_row, column=1, value="總計")
ws.cell(row=summary_row, column=1).font = Font(bold=True)

quoted_total = sum(float(r[4]) * r[3] for r in data if r[4])
ws.cell(row=summary_row, column=5, value=f"{quoted_total:.2f}")
ws.cell(row=summary_row, column=5).font = Font(bold=True)

received_total = sum(float(r[5]) for r in data if r[5])
ws.cell(row=summary_row, column=6, value=f"{received_total:.2f}")
ws.cell(row=summary_row, column=6).font = Font(bold=True)

# Auto-adjust column width
for col_idx in range(1, len(headers) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = 10
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) + 2)
    ws.column_dimensions[col_letter].width = min(max_len, 45)

output_path = "output/dave_transactions_v2.xlsx"
wb.save(output_path)
print(f"Done: {output_path}")
print(f"Records: {len(data)}")
print(f"Quoted total (qty x unit): ${quoted_total:.2f}")
print(f"Received total (confirmed paid): ${received_total:.2f}")
print(f"Outstanding (unpaid quoted): ${quoted_total - received_total:.2f}")
