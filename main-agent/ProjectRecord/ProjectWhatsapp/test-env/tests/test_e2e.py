"""End-to-end integration tests for the full pipeline.

Tests the complete flow: parse_chat_file → build_records → export_to_excel.
Uses real modules for parsing, building, and exporting.
Uses mock ImageAnalysisResult (no real OCR calls).
"""
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.parser.text_parser import parse_chat_file
from src.builder.record_builder import build_records
from src.exporter.excel_exporter import export_to_excel
from src.models.image_result import ImageAnalysisResult


FIXTURES_DIR = Path(__file__).parent / "fixtures" / "e2e"


@pytest.fixture
def sample_chat_path() -> Path:
    path = FIXTURES_DIR / "sample_chat.txt"
    assert path.exists(), f"Fixture not found: {path}"
    return path


@pytest.fixture
def mock_image_results() -> list[ImageAnalysisResult]:
    return [
        ImageAnalysisResult(
            filename="img001.jpg", analysis_mode="ocr",
            payment_method="payme", amount=Decimal("500"),
            transaction_date=date(2024, 1, 15), confidence=0.95,
        ),
        ImageAnalysisResult(
            filename="img002.jpg", analysis_mode="ocr",
            payment_method="fps", amount=Decimal("600"),
            transaction_date=date(2024, 1, 15), confidence=0.90,
        ),
        ImageAnalysisResult(
            filename="img003.jpg", analysis_mode="ocr",
            payment_method="bank_transfer", amount=Decimal("800"),
            transaction_date=date(2024, 1, 15), confidence=0.88,
        ),
        ImageAnalysisResult(
            filename="img004.jpg", analysis_mode="ocr",
            payment_method="payme", amount=Decimal("300"),
            transaction_date=date(2024, 1, 15), confidence=0.92,
        ),
        ImageAnalysisResult(
            filename="img005.jpg", analysis_mode="ocr",
            payment_method="fps", amount=Decimal("1200"),
            transaction_date=date(2024, 1, 15), confidence=0.85,
        ),
    ]


@pytest.fixture
def tmp_output_path(tmp_path) -> Path:
    return tmp_path / "output.xlsx"


class TestFullPipelineHappyPath:
    def test_full_pipeline_happy_path(
        self, sample_chat_path, mock_image_results, tmp_output_path
    ):
        messages = parse_chat_file(str(sample_chat_path))
        records = build_records(messages, mock_image_results)
        export_to_excel(records, str(tmp_output_path))

        wb = load_workbook(tmp_output_path)
        ws = wb.active
        assert _count_data_rows(ws) == 5

    def test_full_pipeline_verify_amounts(
        self, sample_chat_path, mock_image_results, tmp_output_path
    ):
        messages = parse_chat_file(str(sample_chat_path))
        records = build_records(messages, mock_image_results)
        export_to_excel(records, str(tmp_output_path))

        wb = load_workbook(tmp_output_path)
        ws = wb.active
        received_amounts = _get_column_values(ws, col=6)
        assert len(received_amounts) == 5
        amount_set = set(received_amounts)
        assert "500.00" in amount_set
        assert "800.00" in amount_set
        assert "1200.00" in amount_set

    def test_full_pipeline_verify_customers(
        self, sample_chat_path, mock_image_results, tmp_output_path
    ):
        messages = parse_chat_file(str(sample_chat_path))
        records = build_records(messages, mock_image_results)
        export_to_excel(records, str(tmp_output_path))

        wb = load_workbook(tmp_output_path)
        ws = wb.active
        customers = _get_column_values(ws, col=2)
        expected_names = {"陳大文", "李小明", "王美麗", "張三", "趙四"}
        assert set(customers) == expected_names


class TestErrorImageMixed:
    def test_error_image_mixed(
        self, sample_chat_path, mock_image_results, tmp_output_path
    ):
        error_result = ImageAnalysisResult(
            filename="img003.jpg", analysis_mode="ocr",
            payment_method=None, amount=None, confidence=0.3,
            error="OCR failed: image too blurry", needs_review=True,
        )
        modified_results = list(mock_image_results)
        modified_results[2] = error_result

        messages = parse_chat_file(str(sample_chat_path))
        records = build_records(messages, modified_results)
        export_to_excel(records, str(tmp_output_path))

        assert len(records) >= 4
        review_records = [r for r in records if r.needs_review]
        assert len(review_records) >= 1

        wb = load_workbook(tmp_output_path)
        ws = wb.active
        assert _count_data_rows(ws) >= 4


class TestEmptyChatFile:
    def test_empty_chat_file(self, tmp_path, tmp_output_path):
        empty_file = tmp_path / "empty_chat.txt"
        empty_file.write_text("", encoding="utf-8")

        messages = parse_chat_file(str(empty_file))
        assert messages == []

        records = build_records(messages, [])
        assert records == []

        export_to_excel(records, str(tmp_output_path))

        wb = load_workbook(tmp_output_path)
        ws = wb.active
        assert _count_data_rows(ws) == 0


class TestNoTransactionContent:
    def test_no_transaction_content(self, tmp_path, tmp_output_path):
        chat_content = (
            "[2024/01/15, 10:00:00] 陳大文: 你好\n"
            "[2024/01/15, 10:01:00] 陳大文: <attached: img001.jpg>\n"
            "[2024/01/15, 11:00:00] 李小明: 今日天氣好好\n"
            "[2024/01/15, 11:01:00] 李小明: <attached: img002.jpg>\n"
        )
        chat_file = tmp_path / "no_transaction.txt"
        chat_file.write_text(chat_content, encoding="utf-8")

        image_results = [
            ImageAnalysisResult(
                filename="img001.jpg", analysis_mode="ocr",
                payment_method="payme", amount=Decimal("500"), confidence=0.9,
            ),
            ImageAnalysisResult(
                filename="img002.jpg", analysis_mode="ocr",
                payment_method="fps", amount=Decimal("300"), confidence=0.85,
            ),
        ]

        messages = parse_chat_file(str(chat_file))
        records = build_records(messages, image_results)

        assert len(records) >= 1
        review_records = [r for r in records if r.needs_review]
        assert len(review_records) >= 1

        export_to_excel(records, str(tmp_output_path))
        wb = load_workbook(tmp_output_path)
        ws = wb.active
        assert _count_data_rows(ws) >= 1


def _count_data_rows(ws) -> int:
    count = 0
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value is not None and "總計" in str(cell_value):
            break
        if cell_value is not None:
            count += 1
    return count


def _get_column_values(ws, col: int) -> list[str]:
    values = []
    for row_idx in range(2, ws.max_row + 1):
        first_col = ws.cell(row=row_idx, column=1).value
        if first_col is not None and "總計" in str(first_col):
            break
        cell_value = ws.cell(row=row_idx, column=col).value
        if cell_value is not None:
            values.append(str(cell_value))
    return values
