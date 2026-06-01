"""Unit tests for src/exporter/excel_exporter.py."""
import time
import pytest
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from src.exporter.excel_exporter import ExportError, export_to_excel
from src.models.transaction import TransactionRecord


def _make_record(
    customer_name: str = "陳大文",
    transaction_date: date = date(2026, 5, 1),
    **kwargs,
) -> TransactionRecord:
    defaults = {
        "transaction_date": transaction_date,
        "customer_name": customer_name,
        "repair_item": "換屏幕",
        "quoted_amount": Decimal("500.00"),
        "received_amount": Decimal("500.00"),
        "payment_method": "payme",
        "payment_status": "paid",
        "notes": "數量: 1",
        "confidence": 0.9,
        "needs_review": False,
    }
    defaults.update(kwargs)
    return TransactionRecord(**defaults)


def _cell_value_or_empty(ws, row, col):
    val = ws.cell(row=row, column=col).value
    return val if val is not None else ""


class TestExportToExcelHappyPath:
    def test_export_three_records(self, tmp_path):
        records = [
            _make_record("陳大文", date(2026, 5, 3)),
            _make_record("李小明", date(2026, 5, 1)),
            _make_record("王美麗", date(2026, 5, 2)),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        assert ws.max_row == 5
        assert ws.max_column == 9

    def test_sort_by_date_default(self, tmp_path):
        records = [
            _make_record("陳大文", date(2026, 5, 3)),
            _make_record("李小明", date(2026, 5, 1)),
            _make_record("王美麗", date(2026, 5, 2)),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "李小明"
        assert ws.cell(row=3, column=2).value == "王美麗"
        assert ws.cell(row=4, column=2).value == "陳大文"

    def test_summary_row_totals(self, tmp_path):
        records = [
            _make_record(quoted_amount=Decimal("100.00"), received_amount=Decimal("90.00")),
            _make_record(quoted_amount=Decimal("200.00"), received_amount=Decimal("200.00")),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        summary_row = len(records) + 2
        assert ws.cell(row=summary_row, column=5).value == "300.00"
        assert ws.cell(row=summary_row, column=6).value == "290.00"


class TestExportToExcelErrorPath:
    def test_output_dir_not_exist_raises_export_error(self, tmp_path):
        bad_path = str(tmp_path / "nonexistent" / "test.xlsx")
        with pytest.raises(ExportError) as exc_info:
            export_to_excel([_make_record()], bad_path)
        assert "輸出目錄不存在" in str(exc_info.value)

    def test_empty_records_produces_header_only(self, tmp_path):
        output = str(tmp_path / "test.xlsx")
        export_to_excel([], output)

        wb = load_workbook(output)
        ws = wb.active
        assert ws.max_row == 2
        assert ws.cell(row=1, column=1).value == "日期"
        assert ws.cell(row=2, column=1).value == "總計"


class TestExportToExcelEdgeCases:
    def test_customer_name_with_emoji(self, tmp_path):
        records = [_make_record(customer_name="陳大文 🔧")]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "陳大文 🔧"

    def test_amount_none_shows_empty(self, tmp_path):
        records = [_make_record(quoted_amount=None, received_amount=None)]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        assert _cell_value_or_empty(ws, 2, 5) == ""
        assert _cell_value_or_empty(ws, 2, 6) == ""

    def test_performance_100_records(self, tmp_path):
        records = [
            _make_record(
                customer_name=f"客戶{i}",
                quoted_amount=Decimal(str(i * 10)),
                received_amount=Decimal(str(i * 9)),
            )
            for i in range(150)
        ]
        output = str(tmp_path / "test.xlsx")
        start = time.time()
        export_to_excel(records, output)
        elapsed = time.time() - start
        assert elapsed < 3.0
