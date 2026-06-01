"""Unit tests for src/exporter/excel_exporter.py."""
import time
import pytest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src.exporter.excel_exporter import ExportError, export_to_excel
from src.models.transaction import TransactionRecord


def _make_record(
    customer_name: str = "陳大文",
    transaction_date: date = date(2026, 5, 1),
    **kwargs,
) -> TransactionRecord:
    """建立測試用 TransactionRecord。"""
    defaults = {
        "transaction_date": transaction_date,
        "customer_name": customer_name,
        "repair_item": "換屏幕",
        "quoted_amount": Decimal("500.00"),
        "received_amount": Decimal("500.00"),
        "payment_method": "payme",
        "payment_status": "paid",
        "notes": "數量: 1",
        "confidence": 0.9,
        "needs_review": False,
    }
    defaults.update(kwargs)
    return TransactionRecord(**defaults)


def _cell_value_or_empty(ws, row, col):
    """讀取 cell 值，None 視為空字串（openpyxl 行為）。"""
    val = ws.cell(row=row, column=col).value
    return val if val is not None else ""


class TestExportToExcelHappyPath:
    """Happy path tests for export_to_excel."""

    def test_export_three_records(self, tmp_path):
        """3 筆記錄匯出後有 1 行表頭 + 3 行數據 + 1 行總計。"""
        records = [
            _make_record("陳大文", date(2026, 5, 3)),
            _make_record("李小明", date(2026, 5, 1)),
            _make_record("王美麗", date(2026, 5, 2)),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        assert ws.max_row == 5  # 1 header + 3 data + 1 summary
        assert ws.max_column == 9

    def test_sort_by_date_default(self, tmp_path):
        """預設按日期排序。"""
        records = [
            _make_record("陳大文", date(2026, 5, 3)),
            _make_record("李小明", date(2026, 5, 1)),
            _make_record("王美麗", date(2026, 5, 2)),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        # Row 2 should be earliest date (李小明, 5/1)
        assert ws.cell(row=2, column=2).value == "李小明"
        assert ws.cell(row=3, column=2).value == "王美麗"
        assert ws.cell(row=4, column=2).value == "陳大文"

    def test_sort_by_customer_name(self, tmp_path):
        """按客戶名稱排序。"""
        records = [
            _make_record("陳大文", date(2026, 5, 1)),
            _make_record("李小明", date(2026, 5, 2)),
            _make_record("王美麗", date(2026, 5, 3)),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output, sort_by="customer_name")

        wb = load_workbook(output)
        ws = wb.active
        names = [ws.cell(row=r, column=2).value for r in range(2, 5)]
        assert names == sorted(names)

    def test_amount_values_correct(self, tmp_path):
        """金額欄位值正確。"""
        records = [
            _make_record(
                quoted_amount=Decimal("300.50"),
                received_amount=Decimal("280.00"),
            ),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=5).value == "300.50"
        assert ws.cell(row=2, column=6).value == "280.00"

    def test_summary_row_totals(self, tmp_path):
        """總計行正確計算報價總額同實收總額。"""
        records = [
            _make_record(
                quoted_amount=Decimal("100.00"),
                received_amount=Decimal("90.00"),
            ),
            _make_record(
                quoted_amount=Decimal("200.00"),
                received_amount=Decimal("200.00"),
            ),
            _make_record(
                quoted_amount=Decimal("300.00"),
                received_amount=Decimal("250.00"),
            ),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        summary_row = len(records) + 2
        assert ws.cell(row=summary_row, column=5).value == "600.00"
        assert ws.cell(row=summary_row, column=6).value == "540.00"

    def test_header_row_values(self, tmp_path):
        """表頭行包含正確欄位名稱。"""
        records = [_make_record()]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        expected = [
            "日期", "客戶名稱", "維修項目", "數量",
            "報價金額（單價）", "實收金額", "付款方式",
            "付款狀態", "備註",
        ]
        assert headers == expected


class TestExportToExcelErrorPath:
    """Error path tests for export_to_excel."""

    def test_output_dir_not_exist_raises_export_error(self, tmp_path):
        """輸出目錄唔存在時 raise ExportError。"""
        bad_path = str(tmp_path / "nonexistent" / "test.xlsx")
        records = [_make_record()]

        with pytest.raises(ExportError) as exc_info:
            export_to_excel(records, bad_path)

        assert "輸出目錄不存在" in str(exc_info.value)

    def test_empty_records_produces_header_only(self, tmp_path):
        """空列表產出只有表頭 + 總計嘅 Excel。"""
        output = str(tmp_path / "test.xlsx")
        export_to_excel([], output)

        wb = load_workbook(output)
        ws = wb.active
        # 1 header + 0 data + 1 summary
        assert ws.max_row == 2
        assert ws.cell(row=1, column=1).value == "日期"
        assert ws.cell(row=2, column=1).value == "總計"


class TestExportToExcelEdgeCases:
    """Edge case tests for export_to_excel."""

    def test_customer_name_with_emoji(self, tmp_path):
        """客戶名稱含 emoji 時 Excel 正確顯示。"""
        records = [_make_record(customer_name="陳大文 🔧")]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "陳大文 🔧"

    def test_amount_none_shows_empty(self, tmp_path):
        """金額為 None 時欄位顯示為空（openpyxl 讀回 None 或 ""）。"""
        records = [
            _make_record(
                quoted_amount=None,
                received_amount=None,
            ),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        # openpyxl 將空字串讀回為 None
        assert _cell_value_or_empty(ws, 2, 5) == ""
        assert _cell_value_or_empty(ws, 2, 6) == ""

    def test_performance_100_records(self, tmp_path):
        """100+ records 效能 < 3 秒。"""
        records = [
            _make_record(
                customer_name=f"客戶{i}",
                transaction_date=date(2026, 1, 1),
                quoted_amount=Decimal(str(i * 10)),
                received_amount=Decimal(str(i * 9)),
            )
            for i in range(150)
        ]
        output = str(tmp_path / "test.xlsx")

        start = time.time()
        export_to_excel(records, output)
        elapsed = time.time() - start

        assert elapsed < 3.0

    def test_summary_with_mixed_none_amounts(self, tmp_path):
        """部分金額為 None 時總計只計算有值嘅記錄。"""
        records = [
            _make_record(
                quoted_amount=Decimal("100.00"),
                received_amount=None,
            ),
            _make_record(
                quoted_amount=None,
                received_amount=Decimal("200.00"),
            ),
        ]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        summary_row = len(records) + 2
        assert ws.cell(row=summary_row, column=5).value == "100.00"
        assert ws.cell(row=summary_row, column=6).value == "200.00"

    def test_repair_item_none_shows_empty(self, tmp_path):
        """維修項目為 None 時顯示空字串（openpyxl 讀回 None 或 ""）。"""
        records = [_make_record(repair_item=None)]
        output = str(tmp_path / "test.xlsx")
        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        # openpyxl 將空字串讀回為 None
        assert _cell_value_or_empty(ws, 2, 3) == ""
