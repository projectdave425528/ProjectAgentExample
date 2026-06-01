"""Integration test for Excel exporter — 驗證完整匯出流程。"""
import pytest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src.exporter import export_to_excel, ExportError
from src.exporter.formatters import (
    format_amount,
    format_date,
    format_payment_method,
    format_payment_status,
)
from src.models.transaction import TransactionRecord


def _cell_value_or_empty(ws, row, col):
    """讀取 cell 值，None 視為空字串（openpyxl 行為）。"""
    val = ws.cell(row=row, column=col).value
    return val if val is not None else ""


def _build_test_records() -> list[TransactionRecord]:
    """建立完整測試數據集。"""
    return [
        TransactionRecord(
            transaction_date=date(2026, 5, 15),
            customer_name="陳大文",
            repair_item="換屏幕",
            quoted_amount=Decimal("800.00"),
            received_amount=Decimal("800.00"),
            payment_method="payme",
            payment_status="paid",
            notes="數量: 2",
            confidence=0.95,
            needs_review=False,
        ),
        TransactionRecord(
            transaction_date=date(2026, 5, 10),
            customer_name="李小明 🎮",
            repair_item="換電池",
            quoted_amount=Decimal("350.00"),
            received_amount=Decimal("300.00"),
            payment_method="fps",
            payment_status="partial",
            notes="客戶話遲啲補返",
            confidence=0.8,
            needs_review=True,
        ),
        TransactionRecord(
            transaction_date=date(2026, 5, 20),
            customer_name="王美麗",
            repair_item=None,
            quoted_amount=None,
            received_amount=Decimal("150.00"),
            payment_method=None,
            payment_status="paid",
            notes="",
            confidence=0.7,
            needs_review=False,
        ),
    ]


class TestExcelExporterIntegration:
    """Integration tests — 完整流程驗證。"""

    def test_full_export_and_read_back(self, tmp_path):
        """完整匯出後讀取驗證所有欄位。"""
        records = _build_test_records()
        output = str(tmp_path / "integration.xlsx")

        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active

        # 驗證結構：1 header + 3 data + 1 summary
        assert ws.max_row == 5
        assert ws.max_column == 9

        # 驗證排序（按日期，5/10 < 5/15 < 5/20）
        assert ws.cell(row=2, column=1).value == "2026-05-10"
        assert ws.cell(row=3, column=1).value == "2026-05-15"
        assert ws.cell(row=4, column=1).value == "2026-05-20"

        # 驗證 emoji 客戶名稱
        assert ws.cell(row=2, column=2).value == "李小明 🎮"

        # 驗證 None 欄位（openpyxl 讀回空字串為 None）
        assert _cell_value_or_empty(ws, 4, 3) == ""  # repair_item
        assert _cell_value_or_empty(ws, 4, 5) == ""  # quoted_amount

        # 驗證付款方式格式化
        assert ws.cell(row=2, column=7).value == "轉數快"
        assert ws.cell(row=3, column=7).value == "PayMe"
        assert ws.cell(row=4, column=7).value == "未知"

        # 驗證付款狀態格式化
        assert ws.cell(row=2, column=8).value == "部分付款"
        assert ws.cell(row=3, column=8).value == "已付"

        # 驗證數量提取
        assert ws.cell(row=3, column=4).value == 2  # "數量: 2"
        assert ws.cell(row=2, column=4).value == 1  # 冇數量信息

    def test_summary_row_calculation(self, tmp_path):
        """總計行正確計算（跳過 None 值）。"""
        records = _build_test_records()
        output = str(tmp_path / "summary.xlsx")

        export_to_excel(records, output)

        wb = load_workbook(output)
        ws = wb.active
        summary_row = 5

        # quoted: 800 + 350 = 1150 (第三筆 None 唔計)
        assert ws.cell(row=summary_row, column=5).value == "1150.00"
        # received: 800 + 300 + 150 = 1250
        assert ws.cell(row=summary_row, column=6).value == "1250.00"

    def test_customer_name_sort_integration(self, tmp_path):
        """按客戶名稱排序嘅完整流程。"""
        records = _build_test_records()
        output = str(tmp_path / "sorted.xlsx")

        export_to_excel(records, output, sort_by="customer_name")

        wb = load_workbook(output)
        ws = wb.active
        names = [ws.cell(row=r, column=2).value for r in range(2, 5)]
        assert names == sorted(names)

    def test_export_error_message_format(self, tmp_path):
        """ExportError 包含正確嘅目錄路徑。"""
        bad_dir = tmp_path / "does_not_exist"
        bad_path = str(bad_dir / "test.xlsx")

        with pytest.raises(ExportError) as exc_info:
            export_to_excel([], bad_path)

        error_msg = str(exc_info.value)
        assert "輸出目錄不存在" in error_msg
        assert "does_not_exist" in error_msg
