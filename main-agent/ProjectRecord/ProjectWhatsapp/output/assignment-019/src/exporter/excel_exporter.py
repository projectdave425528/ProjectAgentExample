"""Excel 匯出模組 — 將 TransactionRecord 列表寫入 .xlsx 文件。"""
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.exporter.formatters import (
    extract_quantity_from_notes,
    format_amount,
    format_date,
    format_payment_method,
    format_payment_status,
)


class ExportError(Exception):
    """匯出過程中嘅錯誤。"""


HEADERS = [
    "日期",
    "客戶名稱",
    "維修項目",
    "數量",
    "報價金額（單價）",
    "實收金額",
    "付款方式",
    "付款狀態",
    "備註",
]

HEADER_FILL = PatternFill(
    start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
)
HEADER_FONT = Font(bold=True)

MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 30


def export_to_excel(records: list, output_path: str, sort_by: str = "date") -> None:
    """將 TransactionRecord 列表匯出為 .xlsx 文件。

    Args:
        records: TransactionRecord 列表
        output_path: 輸出文件路徑
        sort_by: 排序方式 ("date" 或 "customer_name")

    Raises:
        ExportError: 輸出目錄不存在時
    """
    _validate_output_path(output_path)
    sorted_records = _sort_records(records, sort_by)

    wb = Workbook()
    ws = wb.active
    ws.title = "交易記錄"

    _write_header(ws)
    _write_data_rows(ws, sorted_records)
    _write_summary_row(ws, sorted_records)
    _auto_adjust_column_width(ws)

    wb.save(output_path)


def _validate_output_path(output_path: str) -> None:
    """驗證輸出目錄是否存在。"""
    dir_path = Path(output_path).parent
    if not dir_path.exists():
        raise ExportError(f"輸出目錄不存在: {dir_path}")


def _sort_records(records: list, sort_by: str) -> list:
    """按指定欄位排序記錄。"""
    if sort_by == "customer_name":
        return sorted(records, key=lambda r: r.customer_name)
    return sorted(records, key=lambda r: r.transaction_date)


def _write_header(ws) -> None:
    """寫入表頭行（粗體 + 淺藍背景）。"""
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_data_rows(ws, records: list) -> None:
    """寫入數據行。"""
    for row_idx, record in enumerate(records, start=2):
        row_data = _record_to_row(record)
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _record_to_row(record) -> list:
    """將單筆 TransactionRecord 轉換為一行數據。"""
    return [
        format_date(record.transaction_date),
        record.customer_name,
        record.repair_item or "",
        extract_quantity_from_notes(record.notes),
        format_amount(record.quoted_amount),
        format_amount(record.received_amount),
        format_payment_method(record.payment_method),
        format_payment_status(record.payment_status),
        record.notes,
    ]


def _write_summary_row(ws, records: list) -> None:
    """寫入總計行（報價總額 + 實收總額）。"""
    row_idx = len(records) + 2
    ws.cell(row=row_idx, column=1, value="總計")
    ws.cell(row=row_idx, column=1).font = Font(bold=True)

    quoted_total = _sum_amounts(records, "quoted_amount")
    received_total = _sum_amounts(records, "received_amount")

    ws.cell(row=row_idx, column=5, value=format_amount(quoted_total))
    ws.cell(row=row_idx, column=6, value=format_amount(received_total))


def _sum_amounts(records: list, field: str) -> Decimal | None:
    """計算指定金額欄位嘅總和。"""
    total = Decimal("0")
    has_value = False
    for record in records:
        amount = getattr(record, field, None)
        if amount is not None:
            total += amount
            has_value = True
    return total if has_value else None


def _auto_adjust_column_width(ws) -> None:
    """根據內容自動調整欄寬。"""
    for col_idx in range(1, len(HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = MIN_COL_WIDTH
        for row in ws.iter_rows(
            min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                if cell.value:
                    length = len(str(cell.value)) + 2
                    max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = min(
            max_length, MAX_COL_WIDTH
        )
